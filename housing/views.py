import openpyxl
from openpyxl.styles import PatternFill
from django.contrib import messages
from django.http import HttpResponse
from django.views.generic import TemplateView
from .models import Comunidad, RepresentanteFamilia, ProgresoVivienda
from django.urls import reverse_lazy
from django.views.generic.edit import FormView
from django.views.generic import DetailView, ListView, UpdateView, DeleteView
from .forms import ComunidadForm, RepresentanteFamiliaForm, ProgresoForm
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib.auth.views import LogoutView
from django.template.loader import render_to_string
from xhtml2pdf import pisa

class LogoutView(LogoutView):
    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)



@method_decorator(login_required, name='dispatch')
class IndexView(TemplateView):
    template_name = 'index.html'

@method_decorator(login_required, name='dispatch')
class AboutView(TemplateView):
    template_name = 'about.html'

def export_progress_to_excel(request):
    comunidades = Comunidad.objects.all()

    # Validación: Si no hay comunidades o progresos, no generar el archivo
    if not comunidades.exists():
        messages.error(request, "No hay datos registrados para exportar.")
        return HttpResponse(status=204)  # No content

    # Crear un archivo Excel
    workbook = openpyxl.Workbook()
    del workbook["Sheet"]  # Eliminar la hoja predeterminada "Sheet"

    # Recorrer las comunidades y crear una hoja para cada una
    for comunidad in comunidades:
        sheet = workbook.create_sheet(title=comunidad.nombre)
        sheet.append(["Representante", "Fecha", "Estufa Mejorada", "Agua Potable", "Divisiones",
                      "Piso", "Limpieza", "Muebles", "Comida Protegida", 
                      
                      "Plagas", "Jaulas", "Desparasitación", "Plan de Ahorro", "Piscicultura",
                      "Tubérculos", "Árboles Futales", "Cerco Vivo",

                      "Reciclaje", "Baño Diario", "Lavarse las Manos", "Letrina", "Sumidero",
                      "Escuela", "Capacitaciones", "Vacunación", "Peso", "Chequeo de Embarazo",
                      "Emergencia Embarazo",

                      "Promedio"])

        representantes = comunidad.representantes.all()
        
        if not representantes.exists():  # Si la comunidad no tiene representantes
            sheet.append(["No hay representantes para esta comunidad."])
            continue

        # Recorrer los representantes y agregar filas con los progresos
        for representante in representantes:
            progresos = representante.progresos.all()
            
            if not progresos.exists():  # Si el representante no tiene progresos
                sheet.append([representante.nombre, "Sin progresos registrados"])
                continue

            for progreso in progresos:
                # Estilos de colores según porcentaje
                estufa_fill = color_por_porcentaje(progreso.estufa_mejorada)
                agua_fill = color_por_porcentaje(progreso.agua_potable)
                divisiones_fill = color_por_porcentaje(progreso.divisiones)
                piso_fill = color_por_porcentaje(progreso.divisiones)
                limpieza_fill = color_por_porcentaje(progreso.divisiones)
                muebles_fill = color_por_porcentaje(progreso.divisiones)
                comida_protegida_fill = color_por_porcentaje(progreso.divisiones)

                plagas_fill = color_por_porcentaje(progreso.divisiones)
                jaulas_fill = color_por_porcentaje(progreso.divisiones)
                desparacitacion_fill = color_por_porcentaje(progreso.divisiones)
                ahorro_fill = color_por_porcentaje(progreso.divisiones)
                piscicultura_fill = color_por_porcentaje(progreso.divisiones)
                tuberculos_fill = color_por_porcentaje(progreso.divisiones)
                arboles_frutales_fill = color_por_porcentaje(progreso.divisiones)
                cerco_vivo_fill = color_por_porcentaje(progreso.divisiones)

                reciclaje_fill = color_por_porcentaje(progreso.divisiones)
                banio_diario_fill = color_por_porcentaje(progreso.divisiones)
                lavar_manos_fill = color_por_porcentaje(progreso.divisiones)
                letrina_fill = color_por_porcentaje(progreso.divisiones)
                sumidero_fill = color_por_porcentaje(progreso.divisiones)
                escuela_fill = color_por_porcentaje(progreso.divisiones)
                capacitaciones_fill = color_por_porcentaje(progreso.divisiones)
                vacunacion_fill = color_por_porcentaje(progreso.divisiones)
                peso_fill = color_por_porcentaje(progreso.divisiones)
                chequeo_fill = color_por_porcentaje(progreso.divisiones)
                emergencia_fill = color_por_porcentaje(progreso.divisiones)

                row = [
                    representante.nombre,
                    progreso.fecha_registro,
                    progreso.estufa_mejorada,
                    progreso.agua_potable,
                    progreso.divisiones,
                    progreso.piso,
                    progreso.limpieza,
                    progreso.muebles,
                    progreso.comida_protegida,
                    progreso.plagas,
                    progreso.jaulas,
                    progreso.desparacitacion,
                    progreso.ahorro,
                    progreso.piscicultura,
                    progreso.tuberculos,
                    progreso.arboles_frutales,
                    progreso.cerco_vivo,
                    progreso.reciclaje,
                    progreso.banio_diario,
                    progreso.lavar_manos,
                    progreso.letrina,
                    progreso.sumidero,
                    progreso.escuela,
                    progreso.capacitaciones,
                    progreso.vacunacion,
                    progreso.peso,
                    progreso.chequeo_embarazo,
                    progreso.emergencia_embarazo,
                    progreso.promedio_progreso  # Promedio de progreso
                ]
                sheet.append(row)

                # Aplicar el color en las celdas del progreso
                sheet.cell(row=sheet.max_row, column=3).fill = estufa_fill
                sheet.cell(row=sheet.max_row, column=4).fill = agua_fill
                sheet.cell(row=sheet.max_row, column=5).fill = divisiones_fill
                sheet.cell(row=sheet.max_row, column=6).fill = piso_fill
                sheet.cell(row=sheet.max_row, column=7).fill = limpieza_fill
                sheet.cell(row=sheet.max_row, column=8).fill = muebles_fill
                sheet.cell(row=sheet.max_row, column=9).fill = comida_protegida_fill

                sheet.cell(row=sheet.max_row, column=10).fill = plagas_fill
                sheet.cell(row=sheet.max_row, column=11).fill = jaulas_fill
                sheet.cell(row=sheet.max_row, column=12).fill = desparacitacion_fill
                sheet.cell(row=sheet.max_row, column=13).fill = ahorro_fill
                sheet.cell(row=sheet.max_row, column=14).fill = piscicultura_fill
                sheet.cell(row=sheet.max_row, column=15).fill = tuberculos_fill
                sheet.cell(row=sheet.max_row, column=16).fill = arboles_frutales_fill
                sheet.cell(row=sheet.max_row, column=17).fill = cerco_vivo_fill

                sheet.cell(row=sheet.max_row, column=18).fill = reciclaje_fill
                sheet.cell(row=sheet.max_row, column=19).fill = banio_diario_fill
                sheet.cell(row=sheet.max_row, column=20).fill = lavar_manos_fill
                sheet.cell(row=sheet.max_row, column=21).fill = letrina_fill
                sheet.cell(row=sheet.max_row, column=22).fill = sumidero_fill
                sheet.cell(row=sheet.max_row, column=23).fill = escuela_fill
                sheet.cell(row=sheet.max_row, column=24).fill = capacitaciones_fill
                sheet.cell(row=sheet.max_row, column=25).fill = vacunacion_fill
                sheet.cell(row=sheet.max_row, column=26).fill = peso_fill
                sheet.cell(row=sheet.max_row, column=27).fill = chequeo_fill
                sheet.cell(row=sheet.max_row, column=28).fill = emergencia_fill


    # Guardar el archivo Excel y enviarlo como respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=progresos_vivienda.xlsx'
    workbook.save(response)
    return response

def color_por_porcentaje(porcentaje):
    if porcentaje == 100:
        return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde
    elif porcentaje == 50:
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
    else:
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo
    

def generar_pdf_progreso(request, community_pk, representante_id, pk):
    progreso = ProgresoVivienda.objects.get(pk=pk, representante_id=representante_id, representante__comunidad_id=community_pk)
    html_string = render_to_string('progress/progress_pdf.html', {'progreso': progreso})
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="progreso_{progreso.representante}_{progreso.fecha_registro}_{progreso.pk}.pdf"'

    # Crear PDF con xhtml2pdf
    pisa_status = pisa.CreatePDF(html_string, dest=response)

    if pisa_status.err:
        return HttpResponse('Error al generar PDF', status=500)
    
    return response


# ------------- COMUNIDADES ---------------
@method_decorator(login_required, name='dispatch')
class CommunitiesView(TemplateView):
    template_name = 'communities/communities.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['comunidades'] = Comunidad.objects.all()  
        return context

@method_decorator(login_required, name='dispatch')
class CreateCommunities(FormView):
    template_name = 'communities/create_communities.html'
    form_class = ComunidadForm
    success_url = reverse_lazy('housing:communities')

    def form_valid(self, form):
        form.save()  # Guarda la nueva comunidad
        return super().form_valid(form)
    
@method_decorator(login_required, name='dispatch')
class EditCommunities(UpdateView):
    model = Comunidad
    template_name = 'communities/edit_communities.html'
    form_class = ComunidadForm
    context_object_name = 'comunidad'
    success_url = reverse_lazy('housing:communities')

    def form_valid(self, form):
        return super().form_valid(form)
    
@method_decorator(login_required, name='dispatch')
class DeleteCommunity(DeleteView):
    model = Comunidad
    template_name = 'communities/delete_community.html'
    context_object_name = 'comunidad'
    success_url = reverse_lazy('housing:communities')

    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)


# ------------- PROGRESOS ---------------
@method_decorator(login_required, name='dispatch')
class ProgressCommunitiesView(TemplateView): #lista todas las comunidades
    template_name = 'progress/communities.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['comunidades'] = Comunidad.objects.all()  
        return context

@method_decorator(login_required, name='dispatch')
class ComunidadDetailView(DetailView): #lista todos los representantes de la comunidad seleccionada
    model = Comunidad
    template_name = 'progress/families.html'
    context_object_name = 'comunidad'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['familias'] = self.object.representantes.all()
        return context
    
@method_decorator(login_required, name='dispatch')
class ProgresoViviendaListView(ListView):
    model = ProgresoVivienda
    template_name = 'progress/family_progress.html'
    context_object_name = 'progresos'

    def get_queryset(self):
        representante_id = self.kwargs['representante_id']
        return ProgresoVivienda.objects.filter(representante_id=representante_id)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        community_pk = self.kwargs['community_pk']
        representante_id = self.kwargs['representante_id']

        # Obtener la comunidad y el representante
        comunidad = Comunidad.objects.get(pk=community_pk)
        representante = RepresentanteFamilia.objects.get(pk=representante_id)

        # Añadirlos al contexto
        context['progresos'] = ProgresoVivienda.objects.filter(representante__id=representante_id)
        context['community_pk'] = community_pk
        context['representante_id'] = representante_id
        context['comunidad'] = comunidad
        context['familia'] = representante
        return context
    
@method_decorator(login_required, name='dispatch')
class AgregarProgresoView(FormView):
    template_name = 'progress/create_progress.html'
    form_class = ProgresoForm

    def form_valid(self, form):
        representante_id = self.kwargs['representante_id']
        representante = RepresentanteFamilia.objects.get(id=representante_id)
        progreso = form.save(commit=False)
        progreso.representante = representante  # Asigna el representante automáticamente
        progreso.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['comunidad'] = Comunidad.objects.get(pk=self.kwargs['community_pk'])
        context['familia'] = RepresentanteFamilia.objects.get(id=self.kwargs['representante_id'])
        community_pk = self.kwargs['community_pk']
        representante_id = self.kwargs['representante_id']
        
        context['community_pk'] = community_pk
        context['representante_id'] = representante_id
        return context

    def get_success_url(self):
        return reverse_lazy('housing:progreso_vivienda', kwargs={
            'community_pk': self.kwargs['community_pk'],
            'representante_id': self.kwargs['representante_id']
        })

@method_decorator(login_required, name='dispatch')
class ProgresoViviendaDetailView(DetailView):
    model = ProgresoVivienda
    template_name = 'progress/progress_detail.html'
    context_object_name = 'progreso'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        community_pk = self.kwargs['community_pk']
        representante_id = self.kwargs['representante_id']
        
        context['community_pk'] = community_pk
        context['representante_id'] = representante_id
        return context
    
@method_decorator(login_required, name='dispatch')
class ProgresoViviendaUpdateView(UpdateView):
    model = ProgresoVivienda
    form_class = ProgresoForm
    template_name = 'progress/edit_progress.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtener la comunidad y representante
        community_pk = self.kwargs['community_pk']
        representante_id = self.kwargs['representante_id']
        
        context['community'] = Comunidad.objects.get(pk=community_pk)
        context['community_pk'] = community_pk
        context['representante_id'] = representante_id
        context['familia'] = RepresentanteFamilia.objects.get(pk=representante_id)
        return context
    
    def get_success_url(self):
        return reverse_lazy('housing:progreso_vivienda', kwargs={
            'community_pk': self.kwargs['community_pk'], 
            'representante_id': self.kwargs['representante_id']
        })
    
@method_decorator(login_required, name='dispatch')
class ProgresoViviendaDeleteView(DeleteView):
    model = ProgresoVivienda
    template_name = 'progress/delete_progress.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtener la comunidad y representante
        community_pk = self.kwargs['community_pk']
        representante_id = self.kwargs['representante_id']
        
        context['community'] = Comunidad.objects.get(pk=community_pk)
        context['community_pk'] = community_pk
        context['representante_id'] = representante_id
        context['familia'] = RepresentanteFamilia.objects.get(pk=representante_id)
        return context
    
    def get_success_url(self):
        # Redirigir a la lista de progresos de la familia seleccionada después de la eliminación
        return reverse_lazy('housing:progreso_vivienda', kwargs={
            'community_pk': self.kwargs['community_pk'], 
            'representante_id': self.kwargs['representante_id']
        })
    
# ------------- FAMILIAS ---------------
@method_decorator(login_required, name='dispatch')
class FamiliesView(TemplateView):
    template_name = 'families/families.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['familias'] = RepresentanteFamilia.objects.all()  
        return context

@method_decorator(login_required, name='dispatch')
class CreateFamilies(FormView):
    template_name = 'families/create_families.html'
    form_class = RepresentanteFamiliaForm
    success_url = reverse_lazy('housing:families')

    def form_valid(self, form):
        form.save()  # Guarda la nueva comunidad
        return super().form_valid(form)
    
@method_decorator(login_required, name='dispatch')
class EditFamilies(UpdateView):
    model = RepresentanteFamilia
    template_name = 'families/edit_families.html'
    form_class = RepresentanteFamiliaForm
    context_object_name = 'familia'
    success_url = reverse_lazy('housing:families')

    def form_valid(self, form):
        return super().form_valid(form)
    
@method_decorator(login_required, name='dispatch')
class DeleteFamily(DeleteView):
    model = RepresentanteFamilia
    template_name = 'families/delete_family.html'
    context_object_name = 'familia'
    success_url = reverse_lazy('housing:families')

    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)